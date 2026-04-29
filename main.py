from nicegui import ui
from supabase import create_client, Client
import os
import csv
import io
import calendar
from collections import defaultdict
from datetime import date
from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

MONTHS = [
    (1, 'Jan'), (2, 'Feb'), (3, 'Mar'), (4, 'Apr'), (5, 'May'), (6, 'Jun'),
    (7, 'Jul'), (8, 'Aug'), (9, 'Sep'), (10, 'Oct'), (11, 'Nov'), (12, 'Dec'),
]

# --- SETUP DATABASE ---
load_dotenv() 
url = os.environ.get("SUPABASE_URL")
key = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(url, key)

ui.page_title('Asset Register')

# --- MUJI THEME ENGINE ---
# 1. Override the default vibrant colors with muted Muji tones
ui.colors(
    primary='#7F0019',    # Muji signature deep red
    secondary='#F5F5F5',  # Muted light grey
    accent='#737373',     # Dark grey for secondary text
    positive='#4A6741',   # Muted earthy green
)

# 2. Force the entire app background to a warm, unbleached paper color
ui.query('body').classes('bg-[#FAF9F6] text-[#333333]')

# --- DATA FUNCTIONS ---
def get_assets():
    # Fetching more fields (id, raw category/location IDs, date) to support editing
    response = supabase.table("assets").select("*, categories(name), locations(name)").order("id", desc=True).execute()
    assets = []
    for row in response.data:
        assets.append({
            'id': row['id'],
            'name': row['name'],
            'category': row['categories']['name'] if row['categories'] else 'N/A',
            'category_id': row['category_id'],
            'location': row['locations']['name'] if row['locations'] else 'N/A',
            'location_id': row['location_id'],
            'status': row['status'],
            'price_raw': row['purchase_cost'],
            'price': f"RM {row['purchase_cost']:,.2f}",
            'quantity': row.get('quantity', 1),
            'unit_cost': row.get('unit_cost', 0.0),
            'purchase_date': row['purchase_date'],
            'purchase_year': row.get('purchase_year'),
            'depreciation_rate': row.get('depreciation_rate', 0.0), # Ensure we handle missing depreciation rate
            'remarks': row.get('remarks', '')
        })
    return assets


# --- YEAR-END REPORT GENERATOR ---

COMPANY_NAME = 'MEGA JUTAMAS SDN BHD (663951-U)'

def _resolve_purchase_year(asset):
    py = asset.get('purchase_year')
    if py:
        try:
            return int(py)
        except (ValueError, TypeError):
            pass
    pd = asset.get('purchase_date')
    if pd:
        try:
            return int(str(pd)[:4])
        except (ValueError, TypeError):
            pass
    return None


def _disposal_month(disposal):
    """Best-effort month of disposal (1-12); falls back to 12 (year-end) if unparseable."""
    dd = disposal.get('disposal_date')
    if dd:
        try:
            return int(str(dd).split('-')[1])
        except (ValueError, TypeError, IndexError):
            pass
    return 12


def _disposal_year(disposal):
    dy = disposal.get('disposal_year')
    try:
        return int(dy) if dy else None
    except (ValueError, TypeError):
        return None


def _disposal_acc_dep_at(disposal, original_cost, rate, py):
    """Acc.Dep. allocated to a disposed portion AS AT the time of its disposal."""
    if not original_cost or not py:
        return 0.0
    disp_cost = float(disposal.get('total_disposal_cost') or 0)
    dy = _disposal_year(disposal) or py
    dm = _disposal_month(disposal)
    months_at_disposal = max(0, (dy - py) * 12 + dm)
    monthly = (original_cost * rate) / 12.0
    share = disp_cost / original_cost
    return min(monthly * share * months_at_disposal, disp_cost)


def compute_period_rows(year, month):
    """Per-asset values as at end of the chosen month within the report year."""
    assets = supabase.table('assets').select(
        'id, name, quantity, unit_cost, purchase_cost, depreciation_rate, '
        'purchase_year, purchase_date, status, category_id, location_id, '
        'categories(name), locations(name)'
    ).execute().data
    disposals = supabase.table('disposals').select('*').execute().data

    disp_by_asset = defaultdict(list)
    for d in disposals:
        if d.get('asset_id'):
            disp_by_asset[d['asset_id']].append(d)

    def is_past(d):
        dy = _disposal_year(d)
        return dy is not None and dy < year

    def is_in_period(d):
        dy = _disposal_year(d)
        if dy is None or dy != year:
            return False
        return _disposal_month(d) <= month

    rows = []
    for a in assets:
        py = _resolve_purchase_year(a)
        if py is None or py > year:
            continue

        rate_pct = float(a.get('depreciation_rate') or 0)
        rate = rate_pct / 100.0
        current_cost = float(a.get('purchase_cost') or 0)

        all_disposals = disp_by_asset.get(a['id'], [])
        all_disposed_cost = sum(float(d.get('total_disposal_cost') or 0) for d in all_disposals)
        original_cost = current_cost + all_disposed_cost

        past_disposals = [d for d in all_disposals if is_past(d)]
        period_disposals = [d for d in all_disposals if is_in_period(d)]
        past_disposed_cost = sum(float(d.get('total_disposal_cost') or 0) for d in past_disposals)
        period_disposed_cost = sum(float(d.get('total_disposal_cost') or 0) for d in period_disposals)

        # Cost flow
        if py == year:
            cost_bf = 0.0
            cost_addition = original_cost
        else:
            cost_bf = original_cost - past_disposed_cost
            cost_addition = 0.0
        cost_disposal = period_disposed_cost
        cost_cf = cost_bf + cost_addition - cost_disposal

        # Depreciation: straight-line, allocated by months elapsed
        monthly_charge_full = (original_cost * rate) / 12.0
        months_dep_BF = 0 if py >= year else (year - py) * 12
        months_dep_period_end = (year - py) * 12 + month if py <= year else 0
        acc_dep_full_BF = min(monthly_charge_full * months_dep_BF, original_cost)
        acc_dep_full_at_period = min(monthly_charge_full * months_dep_period_end, original_cost)

        past_disp_acc_dep = sum(_disposal_acc_dep_at(d, original_cost, rate, py) for d in past_disposals)
        period_disp_acc_dep = sum(_disposal_acc_dep_at(d, original_cost, rate, py) for d in period_disposals)

        acc_dep_bf = max(acc_dep_full_BF - past_disp_acc_dep, 0.0)
        acc_dep_disposal = min(period_disp_acc_dep, acc_dep_full_at_period)
        current_charge = max(
            acc_dep_full_at_period - past_disp_acc_dep - period_disp_acc_dep - acc_dep_bf, 0.0
        )
        # Cap so NBV doesn't go negative
        current_charge = min(current_charge, max(cost_cf - (acc_dep_bf - acc_dep_disposal), 0.0))
        monthly_charge = monthly_charge_full
        acc_dep_cf = acc_dep_bf - acc_dep_disposal + current_charge

        nbv_current = cost_cf - acc_dep_cf
        nbv_prior = cost_bf - acc_dep_bf

        rows.append({
            'category': a['categories']['name'] if a.get('categories') else 'Uncategorized',
            'location': a['locations']['name'] if a.get('locations') else 'No location',
            'description': a['name'],
            'reference': '',
            'unit': a.get('quantity') or 0,
            'depreciation_rate': rate,
            'year_of_purchase': py,
            'cost_bf': cost_bf,
            'cost_addition': cost_addition,
            'cost_disposal': cost_disposal,
            'cost_transfer_in': 0.0,
            'cost_transfer_inout': 0.0,
            'cost_cf': cost_cf,
            'acc_dep_bf': acc_dep_bf,
            'acc_dep_disposal': acc_dep_disposal,
            'acc_dep_transfer_in': 0.0,
            'acc_dep_transfer_inout': 0.0,
            'current_charge': current_charge,
            'monthly_charge': monthly_charge,
            'acc_dep_cf': acc_dep_cf,
            'nbv_current': nbv_current,
            'nbv_prior': nbv_prior,
        })
    return rows


def build_period_workbook(year, month):
    rows = compute_period_rows(year, month)
    last_day = calendar.monthrange(year, month)[1]
    period_str = f'{last_day:02d}.{month:02d}.{year}'
    bf_str = f'01.01.{year}'
    prior_str = f'31.12.{year - 1}'

    wb = Workbook()
    ws = wb.active
    month_short = dict(MONTHS).get(month, '').upper()
    ws.title = f'{month_short} {year}'

    bold = Font(bold=True)
    italic = Font(italic=True)
    bold_italic = Font(bold=True, italic=True)
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill('solid', fgColor='F2F2F2')
    thin = Side(style='thin', color='000000')
    border_full = Border(top=thin, bottom=thin, left=thin, right=thin)
    border_vertical = Border(left=thin, right=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

    money_format = '#,##0.00;-#,##0.00;-'
    rate_format = '0.0%'

    LAST_COL = 21  # column U

    def apply_border(rowidx, full=False, start_col=1, end_col=LAST_COL):
        b = border_full if full else border_vertical
        for c in range(start_col, end_col + 1):
            ws.cell(row=rowidx, column=c).border = b

    # Title rows (no border on these)
    ws.cell(row=1, column=1, value=COMPANY_NAME).font = title_font
    ws.cell(row=3, column=1, value=f'FIXED ASSET REGISTERED AS AT {period_str}').font = Font(bold=True, size=11)

    # Column headers (row 5 = group headers, row 6 = sub-headers)
    headers_top = [
        ('A5', 'DESCRIPTIONS'), ('B5', 'REFERENCE'), ('C5', 'UNIT'), ('D5', 'LOCATION'),
        ('E5', 'DEPRECIATION RATE'), ('F5', 'YEAR OF PURCHASE'),
        ('G5', 'COST'),
        ('M5', 'ACCUMULATED DEPRECIATION'),
        ('T5', 'NET BOOK VALUE'),
    ]
    headers_sub = [
        ('G6', f'B/F AS AT\n{bf_str}\nRM'),
        ('H6', 'ADDITION'),
        ('I6', 'DISPOSAL'),
        ('J6', 'TRANSFER IN'),
        ('K6', 'TRANSFER IN/(OUT)'),
        ('L6', f'C/F AS AT\n{period_str}\nRM'),
        ('M6', f'B/F AS AT\n{bf_str}\nRM'),
        ('N6', 'DISPOSAL'),
        ('O6', 'TRANSFER IN'),
        ('P6', 'TRANSFER IN/(OUT)'),
        ('Q6', 'CURRENT CHARGE'),
        ('R6', 'MONTHLY CHARGE'),
        ('S6', f'C/F AS AT\n{period_str}\nRM'),
        ('T6', f'AS AT\n{period_str}\nRM'),
        ('U6', f'AS AT\n{prior_str}\nRM'),
    ]
    for cell, val in headers_top + headers_sub:
        c = ws[cell]
        c.value = val
        c.font = bold
        c.alignment = center
        c.fill = header_fill

    # Merge group cells
    ws.merge_cells('G5:L5')
    ws.merge_cells('M5:S5')
    ws.merge_cells('T5:U5')
    # Description..Year Of Purchase span both header rows
    for col in 'ABCDEF':
        ws.merge_cells(f'{col}5:{col}6')

    # Borders on header rows (full box)
    apply_border(5, full=True)
    apply_border(6, full=True)
    # Row 7 is blank but bridges the headers to the data — keep vertical borders continuous
    apply_border(7)
    ws.row_dimensions[6].height = 50

    # Group rows by category → location
    grouped = defaultdict(lambda: defaultdict(list))
    for r in rows:
        grouped[r['category']][r['location']].append(r)

    SUM_FIELDS = ['cost_bf', 'cost_addition', 'cost_disposal', 'cost_transfer_in', 'cost_transfer_inout',
                  'cost_cf', 'acc_dep_bf', 'acc_dep_disposal', 'acc_dep_transfer_in', 'acc_dep_transfer_inout',
                  'current_charge', 'monthly_charge', 'acc_dep_cf', 'nbv_current', 'nbv_prior']
    FIELD_TO_COL = {
        'cost_bf': 7, 'cost_addition': 8, 'cost_disposal': 9, 'cost_transfer_in': 10,
        'cost_transfer_inout': 11, 'cost_cf': 12, 'acc_dep_bf': 13, 'acc_dep_disposal': 14,
        'acc_dep_transfer_in': 15, 'acc_dep_transfer_inout': 16, 'current_charge': 17,
        'monthly_charge': 18, 'acc_dep_cf': 19, 'nbv_current': 20, 'nbv_prior': 21,
    }

    def write_data_row(rowidx, r):
        ws.cell(row=rowidx, column=1, value=r['description'])
        ws.cell(row=rowidx, column=2, value=r['reference'])
        ws.cell(row=rowidx, column=3, value=r['unit'])
        ws.cell(row=rowidx, column=4, value=r['location'])
        cell_rate = ws.cell(row=rowidx, column=5, value=r['depreciation_rate'])
        cell_rate.number_format = rate_format
        ws.cell(row=rowidx, column=6, value=r['year_of_purchase'])
        for f, col in FIELD_TO_COL.items():
            cell = ws.cell(row=rowidx, column=col, value=round(r[f], 2))
            cell.number_format = money_format
        apply_border(rowidx)

    def write_subtotal_row(rowidx, label, totals, label_font=bold_italic):
        ws.cell(row=rowidx, column=1, value=label).font = label_font
        for f, col in FIELD_TO_COL.items():
            cell = ws.cell(row=rowidx, column=col, value=round(totals[f], 2))
            cell.font = label_font
            cell.number_format = money_format
        apply_border(rowidx)

    grand_totals = {f: 0.0 for f in SUM_FIELDS}
    cur_row = 8

    for category in sorted(grouped.keys()):
        cat_cell = ws.cell(row=cur_row, column=1, value=category)
        cat_cell.font = bold
        apply_border(cur_row)
        cur_row += 1

        cat_totals = {f: 0.0 for f in SUM_FIELDS}

        for location in sorted(grouped[category].keys()):
            loc_cell = ws.cell(row=cur_row, column=4, value=location)
            loc_cell.font = bold
            apply_border(cur_row)
            cur_row += 1

            loc_totals = {f: 0.0 for f in SUM_FIELDS}
            for r in grouped[category][location]:
                write_data_row(cur_row, r)
                for f in SUM_FIELDS:
                    loc_totals[f] += r[f]
                cur_row += 1

            write_subtotal_row(cur_row, f'Subtotal — {location}', loc_totals, italic)
            for f in SUM_FIELDS:
                cat_totals[f] += loc_totals[f]
            cur_row += 1
            apply_border(cur_row)  # keep vertical borders continuous through the blank gap
            cur_row += 1

        write_subtotal_row(cur_row, f'TOTAL {category}', cat_totals, bold)
        for f in SUM_FIELDS:
            grand_totals[f] += cat_totals[f]
        cur_row += 1
        apply_border(cur_row)
        cur_row += 1

    write_subtotal_row(cur_row, 'GRAND TOTAL', grand_totals, bold)

    # Freeze panes: header rows + column A (descriptions) stay visible while scrolling
    ws.freeze_panes = 'B8'

    widths = {'A': 50, 'B': 12, 'C': 6, 'D': 22, 'E': 10, 'F': 8,
              'G': 14, 'H': 12, 'I': 12, 'J': 12, 'K': 14, 'L': 14,
              'M': 14, 'N': 12, 'O': 12, 'P': 14, 'Q': 14, 'R': 14, 'S': 14,
              'T': 14, 'U': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# --- UI LAYOUT ---

# Header: Crisp white background, thin border, zero shadow
with ui.header().classes('bg-[#8b6854] border-b border-stone-200 text-white items-center justify-between px-8 py-3 shadow-none'):
    # A simple, elegant title
    ui.label('Fixed Assets Module Demo').classes('text-xl tracking-wide font-light')
    
    # "Flat" button style for logout to keep it minimal
    ui.button('Sign Out', color='accent').props('flat').classes('text-sm tracking-wider font-light text-white')

# Main Container: Center the content and restrict the width so it doesn't stretch too far on wide monitors
with ui.column().classes('w-full max-w-5xl mx-auto mt-8 px-4'):

    # Tabs: Clean, underlined style rather than heavy colored blocks, aligned to stretch equally
    with ui.tabs().classes('w-full border-b border-stone-200 text-stone-500').props('align="justify"') as tabs:
        tab_dashboard = ui.tab('Dashboard', icon='dashboard')
        tab_add = ui.tab('Addition', icon='add_box')
        tab_disposal = ui.tab('Disposal', icon='delete')
        tab_reports = ui.tab('Reports', icon='bar_chart')
        tab_maintenance = ui.tab('Maintenance', icon='build')

    # Default to Dashboard tab
    with ui.tab_panels(tabs, value=tab_dashboard).classes('w-full bg-transparent p-0 mt-6'):
        
        # --- TAB: DASHBOARD ---
        with ui.tab_panel(tab_dashboard):
            with ui.card().classes('w-full bg-white border border-stone-200 shadow-none rounded-sm p-12 flex flex-col items-center justify-center'):
                ui.icon('dashboard', size='4rem').classes('text-stone-200 mb-4')
                ui.label('Dashboard Overview').classes('text-2xl text-stone-400 font-light mb-2')
                ui.label('Key Performance Indicators and charts will be displayed here.').classes('text-stone-400')

        # --- TAB: ADD ASSET ---
        with ui.tab_panel(tab_add):
            
            # Helper functions to grab the live menus from Supabase
            def get_cat_options():
                try: return {row['id']: row['name'] for row in supabase.table('categories').select('*').execute().data}
                except Exception: return {}

            def get_loc_options():
                try: return {row['id']: row['name'] for row in supabase.table('locations').select('*').execute().data}
                except Exception: return {}

            def get_rate_options():
                try: return {row['percentage']: row['rate_name'] for row in supabase.table('depreciation_rates').select('*').execute().data}
                except Exception: return {}

            # Form Card: Pure white, thin border, sharp corners
            with ui.card().classes('w-full max-w-2xl bg-white border border-stone-200 shadow-none rounded-sm p-8 mx-auto'):
                ui.label('New Asset Details').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-6 w-full')
                
                # Top Row: Text Inputs
                name = ui.input('Asset Description').classes('w-full mb-4').props('outlined dense')
                remarks = ui.input('Remarks').classes('w-full mb-4').props('outlined dense')
                
                # Middle Row: Foreign Key Dropdowns (Categories & Locations)
                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    cat_select = ui.select(get_cat_options(), label='Category').classes('w-1/2').props('outlined dense')
                    loc_select = ui.select(get_loc_options(), label='Location').classes('w-1/2').props('outlined dense')

                # Cost Row: Qty, Unit Cost, and Total Purchase Cost
                def update_total_cost():
                    try:
                        q = float(qty.value) if qty.value else 0.0
                        u = float(unit_cost_input.value) if unit_cost_input.value else 0.0
                        price.value = round(q * u, 2)
                    except (ValueError, TypeError):
                        pass

                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    qty = ui.number('Qty', value=1, min=1, format='%d', on_change=update_total_cost).classes('w-1/4').props('outlined dense')
                    unit_cost_input = ui.number('Unit Cost (RM)', value=0.0, format='%.2f', on_change=update_total_cost).classes('flex-1').props('outlined dense')
                    price = ui.number('Purchase Cost (RM)', format='%.2f').classes('flex-1').props('outlined dense')

                # Rate and Status Row
                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    rate_select = ui.select(get_rate_options(), label='Depreciation Rate', with_input=True).classes('w-1/2').props('outlined dense')
                    status = ui.select(['Active', 'Disposed'], value='Active', label='Status').classes('w-1/2').props('outlined dense')

                # Tiny refresh button so staff can sync the dropdowns if maintenance was updated
                with ui.row().classes('w-full justify-end mb-4'):
                    def refresh_dropdowns():
                        cat_select.options = get_cat_options()
                        cat_select.update()
                        loc_select.options = get_loc_options()
                        loc_select.update()
                        rate_select.options = get_rate_options()
                        rate_select.update()
                        ui.notify('Dropdowns synced with master data', type='info', position='top-right')
                    ui.button('↻ Sync Options', on_click=refresh_dropdowns).props('flat dense').classes('text-stone-400 text-xs tracking-wider')

                # The Save Logic
                def save_asset():
                    # 1. Validation Check
                    if not name.value or qty.value is None or unit_cost_input.value is None:
                        ui.notify('Please fill in the Asset Description, Quantity, and Unit Cost.', type='warning', position='top')
                        return
                    
                    try:
                        # 2. Push to Database
                        supabase.table('assets').insert({
                            'name': name.value,
                            'remarks': remarks.value,
                            'category_id': cat_select.value, # Foreign key link!
                            'location_id': loc_select.value, # Foreign key link!
                            'quantity': int(qty.value) if qty.value else 1,
                            'unit_cost': float(unit_cost_input.value) if unit_cost_input.value else 0.0,
                            'purchase_cost': float(price.value) if price.value else 0.0,
                            'depreciation_rate': float(rate_select.value) if rate_select.value else 0.0,
                            'purchase_date': purchase_date_input.value,
                            'purchase_year': int(purchase_year_input.value) if purchase_year_input.value else None,
                            'status': status.value
                        }).execute()
                        
                        ui.notify('Asset recorded successfully', position='top', type='positive')
                        
                        # 3. Clear the form for the next entry
                        name.value = ''
                        remarks.value = ''
                        qty.value = 1
                        unit_cost_input.value = 0.0
                        price.value = 0.0
                        cat_select.value = None
                        loc_select.value = None
                        rate_select.value = None
                        purchase_date_input.value = str(date.today())
                        purchase_year_input.value = date.today().year
                        
                    except Exception as e:
                        ui.notify(f'Database error: Please check your data or foreign keys. ({e})', type='negative')
                        
                # Primary Button
                ui.button('Save Record', on_click=save_asset).classes('w-full py-2 tracking-wide font-light bg-[#7F0019] text-white').props('unelevated')
                
        # --- TAB: DISPOSAL ---
        with ui.tab_panel(tab_disposal):

            # Pull active assets (with remaining quantity) for the asset selector
            def get_active_assets():
                try:
                    return supabase.table('assets') \
                        .select('id, name, remarks, quantity, unit_cost, purchase_cost, '
                                'category_id, location_id, depreciation_rate, '
                                'purchase_date, purchase_year, categories(name), locations(name)') \
                        .eq('status', 'Active') \
                        .gt('quantity', 0) \
                        .order('name').execute().data
                except Exception:
                    return []

            # Cache so on_asset_select can find the full record from a selected id
            disp_asset_lookup = {}

            def build_disp_asset_options():
                disp_asset_lookup.clear()
                opts = {}
                for a in get_active_assets():
                    disp_asset_lookup[a['id']] = a
                    loc = a['locations']['name'] if a.get('locations') else 'No location'
                    opts[a['id']] = f"{a['name']} ({loc}) — Available: {a.get('quantity') or 0}"
                return opts

            with ui.card().classes('w-full max-w-2xl bg-white border border-stone-200 shadow-none rounded-sm p-8 mx-auto'):
                ui.label('Asset Disposal Details').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-6 w-full')

                # Step 1: pick the asset to dispose
                disp_asset_select = ui.select(
                    build_disp_asset_options(),
                    label='Select Asset to Dispose',
                    with_input=True,
                ).classes('w-full mb-4').props('outlined dense')

                # Auto-populated read-only context fields
                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    disp_category_lbl = ui.input('Category', value='').classes('w-1/2').props('outlined dense readonly')
                    disp_location_lbl = ui.input('Location', value='').classes('w-1/2').props('outlined dense readonly')

                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    disp_available_lbl = ui.number('Available Qty', value=0, format='%d').classes('w-1/2').props('outlined dense readonly')
                    disp_orig_unit_cost_lbl = ui.number('Original Unit Cost (RM)', value=0.0, format='%.2f').classes('w-1/2').props('outlined dense readonly')

                ui.label('Disposal Details').classes('text-sm text-stone-500 mt-2 mb-1 w-full')
                disp_remarks = ui.input('Remarks').classes('w-full mb-4').props('outlined dense')

                def update_total_disp_cost():
                    try:
                        q = float(disp_qty.value) if disp_qty.value else 0.0
                        u = float(disp_unit_cost.value) if disp_unit_cost.value else 0.0
                        total_disp_cost.value = round(q * u, 2)
                    except (ValueError, TypeError):
                        pass

                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    disp_qty = ui.number('Qty Disposed', value=1, min=1, format='%d', on_change=update_total_disp_cost).classes('w-1/3').props('outlined dense')
                    disp_unit_cost = ui.number('Cost per Unit (RM)', value=0.0, format='%.2f', on_change=update_total_disp_cost).classes('w-1/3').props('outlined dense')
                    total_disp_cost = ui.number('Total Cost Disposed (RM)', value=0.0, format='%.2f').classes('w-1/3').props('outlined dense readonly')

                with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                    disp_sales_proceed = ui.number('Sales Proceed (RM)', value=0.0, format='%.2f').classes('w-1/2').props('outlined dense')
                    disp_year_input = ui.number('Year of Disposal', value=date.today().year, format='%d').classes('w-1/2').props('outlined dense readonly')

                with ui.row().classes('w-full mb-4'):
                    def update_disp_year_from_date():
                        try:
                            disp_year_input.value = int(disp_date_input.value[:4])
                        except (ValueError, TypeError):
                            pass

                    disp_date_input = ui.input('Date of Disposal', value=str(date.today()), on_change=lambda: update_disp_year_from_date()).classes('w-full').props('outlined dense')
                    with disp_date_input:
                        with ui.menu().props('no-parent-event') as disp_date_menu:
                            with ui.date().bind_value(disp_date_input) as disp_date_picker:
                                with ui.row().classes('justify-end'):
                                    ui.button('Close', on_click=disp_date_menu.close).props('flat')
                        with disp_date_input.add_slot('append'):
                            ui.icon('edit_calendar').on('click', disp_date_menu.open).classes('cursor-pointer')

                def on_disp_asset_select():
                    aid = disp_asset_select.value
                    if aid is None or aid not in disp_asset_lookup:
                        disp_category_lbl.value = ''
                        disp_location_lbl.value = ''
                        disp_available_lbl.value = 0
                        disp_orig_unit_cost_lbl.value = 0.0
                        return
                    a = disp_asset_lookup[aid]
                    disp_category_lbl.value = a['categories']['name'] if a.get('categories') else ''
                    disp_location_lbl.value = a['locations']['name'] if a.get('locations') else ''
                    disp_available_lbl.value = a.get('quantity') or 0
                    disp_orig_unit_cost_lbl.value = float(a.get('unit_cost') or 0.0)
                    # Default the disposal unit cost to the asset's unit cost
                    disp_unit_cost.value = float(a.get('unit_cost') or 0.0)
                    # Cap qty disposed at the available qty
                    if (disp_qty.value or 0) > (a.get('quantity') or 0):
                        disp_qty.value = a.get('quantity') or 1
                    update_total_disp_cost()

                disp_asset_select.on_value_change(lambda e: on_disp_asset_select())

                with ui.row().classes('w-full justify-end mb-4'):
                    def refresh_disp_asset_options():
                        disp_asset_select.options = build_disp_asset_options()
                        disp_asset_select.update()
                        ui.notify('Asset list refreshed', type='info', position='top-right')
                    ui.button('↻ Refresh Assets', on_click=refresh_disp_asset_options).props('flat dense').classes('text-stone-400 text-xs tracking-wider')

                def save_disposal():
                    aid = disp_asset_select.value
                    if aid is None or aid not in disp_asset_lookup:
                        ui.notify('Please select an asset to dispose.', type='warning', position='top')
                        return

                    asset = disp_asset_lookup[aid]
                    qty_d = int(disp_qty.value) if disp_qty.value else 0
                    available = asset.get('quantity') or 0

                    if qty_d < 1:
                        ui.notify('Quantity disposed must be at least 1.', type='warning', position='top')
                        return
                    if qty_d > available:
                        ui.notify(f'Cannot dispose {qty_d} — only {available} available.', type='warning', position='top')
                        return

                    try:
                        total_cost = float(total_disp_cost.value) if total_disp_cost.value else 0.0

                        # 1. Insert disposal record (linked to the source asset)
                        supabase.table('disposals').insert({
                            'asset_id': aid,
                            'name': asset['name'],
                            'remarks': disp_remarks.value,
                            'category_id': asset.get('category_id'),
                            'location_id': asset.get('location_id'),
                            'sales_proceed': float(disp_sales_proceed.value) if disp_sales_proceed.value else 0.0,
                            'quantity_disposed': qty_d,
                            'unit_cost': float(disp_unit_cost.value) if disp_unit_cost.value else 0.0,
                            'total_disposal_cost': total_cost,
                            'disposal_date': disp_date_input.value,
                            'disposal_year': int(disp_year_input.value) if disp_year_input.value else None,
                            'status': 'Disposed'
                        }).execute()

                        # 2. Decrement source asset; flip status to Disposed when fully gone
                        new_qty = available - qty_d
                        new_purchase_cost = max(float(asset.get('purchase_cost') or 0.0) - total_cost, 0.0)
                        update_payload = {
                            'quantity': new_qty,
                            'purchase_cost': new_purchase_cost,
                        }
                        if new_qty <= 0:
                            update_payload['status'] = 'Disposed'
                        supabase.table('assets').update(update_payload).eq('id', aid).execute()

                        if new_qty <= 0:
                            ui.notify('Asset fully disposed — status set to Disposed.', position='top', type='positive')
                        else:
                            ui.notify(f'Disposal recorded. Remaining quantity: {new_qty}.', position='top', type='positive')

                        # 3. Refresh the asset selector and Reports tab
                        refresh_disp_asset_options()
                        try:
                            refresh_reports()
                        except Exception:
                            pass

                        # 4. Clear the form for the next entry
                        disp_asset_select.value = None
                        disp_remarks.value = ''
                        disp_sales_proceed.value = 0.0
                        disp_qty.value = 1
                        disp_unit_cost.value = 0.0
                        total_disp_cost.value = 0.0
                        disp_category_lbl.value = ''
                        disp_location_lbl.value = ''
                        disp_available_lbl.value = 0
                        disp_orig_unit_cost_lbl.value = 0.0
                        disp_date_input.value = str(date.today())
                        disp_year_input.value = date.today().year

                    except Exception as e:
                        ui.notify(f'Database error: {e}', type='negative')

                ui.button('Record Disposal', on_click=save_disposal).classes('w-full py-2 tracking-wide font-light bg-[#7F0019] text-white').props('unelevated')

        # --- TAB: REPORTS (Including Asset Register) ---
        with ui.tab_panel(tab_reports):
            
            # Header Row with Title and Export Controls
            with ui.row().classes('w-full items-center justify-between mb-4 px-2'):
                ui.label('Comprehensive Asset Register').classes('text-xl text-stone-800 font-light')

                month_options = {m: f'{m:02d} - {label}' for m, label in MONTHS}

                def export_period():
                    try:
                        y = int(year_input.value or date.today().year)
                    except (ValueError, TypeError):
                        y = date.today().year
                    m = month_select.value or date.today().month
                    try:
                        data = build_period_workbook(y, m)
                        last_day = calendar.monthrange(y, m)[1]
                        fname = f'Fixed_Asset_Register_{y}_{m:02d}.xlsx'
                        ui.download(data, fname)
                        ui.notify(f'Report as at {last_day:02d}.{m:02d}.{y} ready', type='positive')
                    except Exception as e:
                        ui.notify(f'Could not build report: {e}', type='negative')

                with ui.row().classes('items-center gap-2'):
                    year_input = ui.number(
                        label='Year', value=date.today().year, format='%d', min=2000, max=2100
                    ).classes('w-24').props('outlined dense')
                    month_select = ui.select(
                        month_options, label='Month', value=date.today().month
                    ).classes('w-32').props('outlined dense')
                    ui.button('Generate Report', icon='description', on_click=export_period) \
                        .classes('text-sm tracking-wide bg-[#7F0019] text-white').props('unelevated')

            # Search Row: Muji-style search bar
            with ui.row().classes('w-full mb-4 px-2'):
                search_input = ui.input(placeholder='Search assets by name, category, or status...').classes('w-full').props('outlined dense clearable')
                search_input.on('input', lambda: asset_table.update()) # Force table update on search

            # The Table: .props('flat bordered') removes the shadow and adds a crisp line
            columns = [
                {'name': 'name', 'label': 'Description', 'field': 'name', 'align': 'left', 'sortable': True},
                {'name': 'category', 'label': 'Category', 'field': 'category', 'align': 'left', 'sortable': True},
                {'name': 'location', 'label': 'Location', 'field': 'location', 'align': 'left', 'sortable': True},
                {'name': 'status', 'label': 'Status', 'field': 'status', 'align': 'center', 'sortable': True},
                {'name': 'price', 'label': 'Purchase Cost', 'field': 'price', 'align': 'right', 'sortable': True},
                {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'},
            ]
            
            asset_data = get_assets()
            asset_table = ui.table(columns=columns, rows=asset_data, row_key='id', pagination=10).classes('w-full text-stone-800').props('flat bordered')
            asset_table.bind_filter_from(search_input, 'value')

            # --- EDIT LOGIC ---
            asset_table.add_slot('body-cell-actions', '''
                <q-td :props="props">
                    <q-btn flat dense icon="edit" color="primary" @click="() => $parent.$emit('edit', props.row)" />
                </q-td>
            ''')

            def refresh_reports():
                asset_table.rows = get_assets()
                asset_table.update()

            def edit_asset(row):
                try:
                    # Pre-load dropdown options before building the dialog
                    try:
                        cat_opts = {r['id']: r['name'] for r in supabase.table('categories').select('*').execute().data}
                    except Exception:
                        cat_opts = {}
                    try:
                        loc_opts = {r['id']: r['name'] for r in supabase.table('locations').select('*').execute().data}
                    except Exception:
                        loc_opts = {}
                    try:
                        rate_opts = {r['percentage']: r['rate_name'] for r in supabase.table('depreciation_rates').select('*').execute().data}
                    except Exception:
                        rate_opts = {}

                    # Build the Edit Dialog — mirrors the Add Assets tab exactly
                    with ui.dialog() as d, ui.card().classes('w-full max-w-2xl bg-white border border-stone-200 shadow-none rounded-sm p-8 mx-auto'):
                        ui.label(f'Edit Asset: {row.get("name", "")}').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-6 w-full')

                        # Top Row: Text Inputs
                        e_name = ui.input('Asset Description', value=row.get('name', '')).classes('w-full mb-4').props('outlined dense')
                        e_remarks = ui.input('Remarks', value=row.get('remarks', '')).classes('w-full mb-4').props('outlined dense')

                        # Middle Row: Foreign Key Dropdowns (Categories & Locations)
                        with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                            cat_val = row.get('category_id')
                            loc_val = row.get('location_id')
                            e_cat = ui.select(cat_opts, label='Category', value=cat_val if cat_val in cat_opts else None).classes('w-1/2').props('outlined dense')
                            e_loc = ui.select(loc_opts, label='Location', value=loc_val if loc_val in loc_opts else None).classes('w-1/2').props('outlined dense')

                        # Cost Row: Qty, Unit Cost, and Total Purchase Cost
                        def update_edit_total_cost():
                            try:
                                q = float(e_qty.value) if e_qty.value else 0.0
                                u = float(e_unit_cost.value) if e_unit_cost.value else 0.0
                                e_price.value = round(q * u, 2)
                            except (ValueError, TypeError):
                                pass

                        with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                            e_qty = ui.number('Qty', value=row.get('quantity', 1), min=1, format='%d', on_change=update_edit_total_cost).classes('w-1/4').props('outlined dense')
                            e_unit_cost = ui.number('Unit Cost (RM)', value=row.get('unit_cost', 0.0), format='%.2f', on_change=update_edit_total_cost).classes('flex-1').props('outlined dense')
                            e_price = ui.number('Purchase Cost (RM)', value=row.get('price_raw'), format='%.2f').classes('flex-1').props('outlined dense')

                        # Rate and Status Row
                        with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                            rate_val = row.get('depreciation_rate', 0.0)
                            e_rate = ui.select(rate_opts, label='Depreciation Rate', value=rate_val if rate_val in rate_opts else None, with_input=True).classes('w-1/2').props('outlined dense')
                            e_status = ui.select(['Active', 'Disposed'], value=row.get('status', 'Active'), label='Status').classes('w-1/2').props('outlined dense')

                        # Date Row: Purchase Date & Year of Purchase
                        with ui.row().classes('w-full gap-4 mb-4 flex-nowrap'):
                            existing_year = row.get('purchase_year') or date.today().year
                            try:
                                existing_year = int(row.get('purchase_date', '')[:4])
                            except (ValueError, TypeError):
                                pass
                            e_purchase_year = ui.number('Year of Purchase', value=existing_year, format='%d').classes('w-1/2').props('outlined dense readonly')

                            def update_edit_year_from_date():
                                try:
                                    e_purchase_year.value = int(e_purchase_date.value[:4])
                                except (ValueError, TypeError):
                                    pass

                            e_purchase_date = ui.input('Date of Purchase', value=row.get('purchase_date', str(date.today())), on_change=lambda: update_edit_year_from_date()).classes('w-1/2').props('outlined dense')
                            with e_purchase_date:
                                with ui.menu().props('no-parent-event') as e_date_menu:
                                    with ui.date().bind_value(e_purchase_date) as e_date_picker:
                                        with ui.row().classes('justify-end'):
                                            ui.button('Close', on_click=e_date_menu.close).props('flat')
                                with e_purchase_date.add_slot('append'):
                                    ui.icon('edit_calendar').on('click', e_date_menu.open).classes('cursor-pointer')

                        # Sync Options button to refresh dropdowns
                        with ui.row().classes('w-full justify-end mb-4'):
                            def refresh_edit_dropdowns():
                                try:
                                    e_cat.options = {r['id']: r['name'] for r in supabase.table('categories').select('*').execute().data}
                                    e_cat.update()
                                    e_loc.options = {r['id']: r['name'] for r in supabase.table('locations').select('*').execute().data}
                                    e_loc.update()
                                    e_rate.options = {r['percentage']: r['rate_name'] for r in supabase.table('depreciation_rates').select('*').execute().data}
                                    e_rate.update()
                                    ui.notify('Dropdowns synced with master data', type='info', position='top-right')
                                except Exception:
                                    ui.notify('Failed to sync dropdowns', type='warning')
                            ui.button('↻ Sync Options', on_click=refresh_edit_dropdowns).props('flat dense').classes('text-stone-400 text-xs tracking-wider')

                        # The Update Logic
                        def save_changes():
                            # 1. Validation Check
                            if not e_name.value or e_qty.value is None or e_unit_cost.value is None:
                                ui.notify('Please fill in the Asset Description, Quantity, and Unit Cost.', type='warning', position='top')
                                return
                            try:
                                # 2. Push update to Database
                                supabase.table('assets').update({
                                    'name': e_name.value,
                                    'remarks': e_remarks.value,
                                    'category_id': e_cat.value,
                                    'location_id': e_loc.value,
                                    'quantity': int(e_qty.value) if e_qty.value else 1,
                                    'unit_cost': float(e_unit_cost.value) if e_unit_cost.value else 0.0,
                                    'purchase_cost': float(e_price.value) if e_price.value else 0.0,
                                    'depreciation_rate': float(e_rate.value) if e_rate.value else 0.0,
                                    'purchase_date': e_purchase_date.value,
                                    'purchase_year': int(e_purchase_year.value) if e_purchase_year.value else None,
                                    'status': e_status.value
                                }).eq('id', row['id']).execute()

                                ui.notify('Asset updated successfully', position='top', type='positive')
                                refresh_reports()
                                d.close()
                            except Exception as e:
                                ui.notify(f'Database error: Please check your data or foreign keys. ({e})', type='negative')

                        # Delete Logic
                        def delete_record():
                            with ui.dialog() as confirm_dialog, ui.card().classes('p-6 bg-white border border-stone-200 shadow-none rounded-sm'):
                                ui.label('Confirm Deletion').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-4 w-full')
                                ui.label(f'Are you sure you want to delete "{row.get("name", "this asset")}"? This action cannot be undone.').classes('text-stone-600 mb-6')
                                with ui.row().classes('w-full gap-4 justify-end'):
                                    ui.button('Cancel', on_click=confirm_dialog.close).props('flat').classes('text-stone-500')
                                    def confirm_delete():
                                        try:
                                            supabase.table('assets').delete().eq('id', row['id']).execute()
                                            ui.notify('Record deleted successfully', type='positive', position='top')
                                            refresh_reports()
                                            confirm_dialog.close()
                                            d.close()
                                        except Exception as e:
                                            ui.notify(f'Delete error: {e}', type='negative')
                                    ui.button('Delete', on_click=confirm_delete).classes('bg-red-700 text-white').props('unelevated')
                            confirm_dialog.open()

                        # Action Buttons
                        ui.button('Update Record', on_click=save_changes).classes('w-full py-2 tracking-wide font-light bg-[#7F0019] text-white').props('unelevated')
                        ui.button('Delete Record', icon='delete', on_click=delete_record).classes('w-full py-2 mt-2 tracking-wide font-light text-red-700').props('flat')
                    d.open()

                except Exception as e:
                    ui.notify(f'Could not open edit dialog: {e}', type='negative')

            asset_table.on('edit', lambda msg: edit_asset(msg.args))

        # --- TAB: MAINTENANCE ---
        with ui.tab_panel(tab_maintenance):
            ui.label('Master Data Management').classes('text-2xl text-stone-800 font-light mb-6')

            with ui.grid(columns=2).classes('w-full gap-6'):
                # --------------------- CATEGORIES ---------------------
                with ui.card().classes('w-full h-full bg-white border border-stone-200 shadow-none rounded-sm p-6 justify-between'):
                    ui.label('Asset Categories').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-4 w-full')
                    
                    cat_cols = [
                        {'name': 'name', 'label': 'Category Name', 'field': 'name', 'align': 'left'},
                        {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'}
                    ]
                    cat_table = ui.table(columns=cat_cols, rows=[], row_key='id', pagination=5).classes('w-full text-stone-800').props('flat bordered dense').style('height: 260px')
                    
                    cat_table.add_slot('body-cell-actions', '''
                        <q-td :props="props">
                            <q-btn flat dense icon="edit" color="primary" @click="() => $parent.$emit('edit', props.row)" />
                            <q-btn flat dense icon="delete" color="negative" @click="() => $parent.$emit('delete', props.row)" />
                        </q-td>
                    ''')

                    def load_categories():
                        try:
                            cat_table.rows = supabase.table('categories').select('*').order('id').execute().data
                        except Exception:
                            pass
                    
                    def del_cat(row):
                        supabase.table('categories').delete().eq('id', row['id']).execute()
                        load_categories()
                        ui.notify('Deleted Category', type='warning')
                        
                    def edit_cat(row):
                        with ui.dialog() as d, ui.card().classes('p-6 min-w-[300px]'):
                            ui.label('Edit Category').classes('text-lg font-bold mb-4')
                            n = ui.input('Name', value=row['name']).classes('w-full mb-4')
                            def save():
                                supabase.table('categories').update({'name': n.value}).eq('id', row['id']).execute()
                                load_categories()
                                d.close()
                                ui.notify('Saved', type='positive')
                            ui.button('Save', on_click=save).classes('w-full bg-[#7F0019] text-white').props('unelevated')
                        d.open()
                        
                    cat_table.on('delete', lambda msg: del_cat(msg.args))
                    cat_table.on('edit', lambda msg: edit_cat(msg.args))
                    load_categories()

                    with ui.row().classes('w-full mt-4 items-center gap-2 flex-nowrap'):
                        new_cat_name = ui.input('New Category').props('outlined dense').classes('flex-1 min-w-[100px]')
                        def add_cat():
                            if new_cat_name.value:
                                supabase.table('categories').insert({'name': new_cat_name.value}).execute()
                                load_categories()
                                new_cat_name.value = ''
                                ui.notify('Added Category', type='positive')
                        ui.button(icon='add', on_click=add_cat).classes('bg-[#7F0019] text-white').props('unelevated round dense')

                # --------------------- LOCATIONS ---------------------
                with ui.card().classes('w-full h-full bg-white border border-stone-200 shadow-none rounded-sm p-6 justify-between'):
                    ui.label('Asset Locations').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-4 w-full')
                    
                    loc_cols = [
                        {'name': 'name', 'label': 'Location Name', 'field': 'name', 'align': 'left'},
                        {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'}
                    ]
                    loc_table = ui.table(columns=loc_cols, rows=[], row_key='id', pagination=5).classes('w-full text-stone-800').props('flat bordered dense').style('height: 260px')
                    loc_table.add_slot('body-cell-actions', '''
                        <q-td :props="props">
                            <q-btn flat dense icon="edit" color="primary" @click="() => $parent.$emit('edit', props.row)" />
                            <q-btn flat dense icon="delete" color="negative" @click="() => $parent.$emit('delete', props.row)" />
                        </q-td>
                    ''')

                    def load_locations():
                        try:
                            loc_table.rows = supabase.table('locations').select('*').order('id').execute().data
                        except Exception:
                            pass

                    def del_loc(row):
                        supabase.table('locations').delete().eq('id', row['id']).execute()
                        load_locations()
                        ui.notify('Deleted Location', type='warning')
                        
                    def edit_loc(row):
                        with ui.dialog() as d, ui.card().classes('p-6 min-w-[300px]'):
                            ui.label('Edit Location').classes('text-lg font-bold mb-4')
                            n = ui.input('Name', value=row['name']).classes('w-full mb-4')
                            def save():
                                supabase.table('locations').update({'name': n.value}).eq('id', row['id']).execute()
                                load_locations()
                                d.close()
                                ui.notify('Saved', type='positive')
                            ui.button('Save', on_click=save).classes('w-full bg-[#7F0019] text-white').props('unelevated')
                        d.open()
                        
                    loc_table.on('delete', lambda msg: del_loc(msg.args))
                    loc_table.on('edit', lambda msg: edit_loc(msg.args))
                    load_locations()

                    with ui.row().classes('w-full mt-4 items-center gap-2 flex-nowrap'):
                        new_loc_name = ui.input('New Location').props('outlined dense').classes('flex-1 min-w-[100px]')
                        def add_loc():
                            if new_loc_name.value:
                                supabase.table('locations').insert({'name': new_loc_name.value}).execute()
                                load_locations()
                                new_loc_name.value = ''
                                ui.notify('Added Location', type='positive')
                        ui.button(icon='add', on_click=add_loc).classes('bg-[#7F0019] text-white').props('unelevated round dense')

            with ui.row().classes('w-full mt-6'):
                # --------------------- DEPRECIATION RATES ---------------------
                with ui.card().classes('w-full bg-white border border-stone-200 shadow-none rounded-sm p-6'):
                    ui.label('Depreciation Rates').classes('text-lg tracking-wide text-stone-800 border-b border-stone-100 pb-2 mb-4 w-full')
                    
                    dep_msg = ui.label('Connecting...').classes('text-sm text-stone-500')
                    
                    dep_cols = [
                        {'name': 'rate_name', 'label': 'Rate Name', 'field': 'rate_name', 'align': 'left'},
                        {'name': 'percentage', 'label': 'Percentage (%)', 'field': 'percentage', 'align': 'center'},
                        {'name': 'actions', 'label': 'Actions', 'field': 'actions', 'align': 'center'}
                    ]
                    dep_table = ui.table(columns=dep_cols, rows=[], row_key='id', pagination=5).classes('w-full text-stone-800').props('flat bordered dense')
                    dep_table.add_slot('body-cell-actions', '''
                        <q-td :props="props">
                            <q-btn flat dense icon="edit" color="primary" @click="() => $parent.$emit('edit', props.row)" />
                            <q-btn flat dense icon="delete" color="negative" @click="() => $parent.$emit('delete', props.row)" />
                        </q-td>
                    ''')

                    def load_dep():
                        try:
                            dep_table.rows = supabase.table('depreciation_rates').select('*').order('id').execute().data
                            dep_msg.set_visibility(False)
                        except Exception as e:
                            dep_msg.set_visibility(True)
                            dep_msg.set_text('Note: Table "depreciation_rates" missing or schema mismatch (needs: id, rate_name, percentage).')
                            dep_msg.classes('text-red-500 mb-2')
                            
                    def del_dep(row):
                        supabase.table('depreciation_rates').delete().eq('id', row['id']).execute()
                        load_dep()
                        ui.notify('Deleted Rate', type='warning')
                        
                    def edit_dep(row):
                        with ui.dialog() as d, ui.card().classes('p-6 min-w-[300px]'):
                            ui.label('Edit Rate').classes('text-lg font-bold mb-4')
                            n = ui.input('Rate Name', value=row['rate_name']).classes('w-full mb-2')
                            p = ui.number('Percentage', value=row['percentage']).classes('w-full mb-4')
                            def save():
                                supabase.table('depreciation_rates').update({'rate_name': n.value, 'percentage': p.value}).eq('id', row['id']).execute()
                                load_dep()
                                d.close()
                                ui.notify('Saved', type='positive')
                            ui.button('Save', on_click=save).classes('w-full bg-[#7F0019] text-white').props('unelevated')
                        d.open()

                    dep_table.on('delete', lambda msg: del_dep(msg.args))
                    dep_table.on('edit', lambda msg: edit_dep(msg.args))
                    load_dep()

                    with ui.row().classes('w-full mt-4 items-center gap-2 flex-nowrap'):
                        new_dep_name = ui.input('New Rate Name').props('outlined dense').classes('flex-1')
                        new_dep_pct = ui.number('Rate (%)', value=10.0).props('outlined dense').classes('w-24')
                        def add_dep():
                            if new_dep_name.value:
                                try:
                                    supabase.table('depreciation_rates').insert({'rate_name': new_dep_name.value, 'percentage': new_dep_pct.value}).execute()
                                    load_dep()
                                    new_dep_name.value = ''
                                    ui.notify('Added Rate', type='positive')
                                except Exception as e:
                                    ui.notify(f"Could not save: {e}", type="negative")
                        ui.button(icon='add', on_click=add_dep).classes('bg-[#7F0019] text-white').props('unelevated round dense')

# --- RUN THE APP ---
port = int(os.environ.get("PORT", 8080))
ui.run(title="Asset App", host="0.0.0.0", port=port)